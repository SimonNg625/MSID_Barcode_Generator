import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import barcode
from barcode.writer import ImageWriter
import re
import os
import io
from pathlib import Path
import tkinter as tk
from tkinter import filedialog

def generate_barcode_image(msid_value):
    """
    Generates a Code128 barcode image in memory.
    """
    Code128 = barcode.get_barcode_class('code128')
    writer = ImageWriter()
    rv = io.BytesIO()
    
    # Configure for a barcode that handles scaling well
    code_instance = Code128(msid_value, writer=writer)
    code_instance.write(rv, options={
        'module_width': 0.3,    
        'module_height': 8.0,   
        'font_size': 6,         
        'text_distance': 3.0,
        'quiet_zone': 1.0,
        'write_text': True
    })
    rv.seek(0)
    return rv

def get_downloads_folder():
    return str(Path.home() / "Downloads")

def process_msid_sheet():
    root = tk.Tk()
    root.withdraw()

    print("Please select your MSID Excel (.xlsx) file...")
    file_path = filedialog.askopenfilename(
        title="Select MSID Excel File",
        filetypes=[("Excel files", "*.xlsx")]
    )

    if not file_path:
        print("No file selected.")
        return

    print(f"Processing file: {file_path}...")

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        msid_pattern = re.compile(r'MSID:?\s*(\d+)', re.IGNORECASE)

        # --- SIZE CONFIGURATION (1.7x Larger) ---
        # Image Size in Pixels
        IMG_WIDTH_PX = 160   # 110 * 1.7 approx
        IMG_HEIGHT_PX = 85   # 50 * 1.7 approx

        # Cell Size Configuration
        # We make the cell slightly larger than the image to create a "frame"
        # 1 char width ≈ 7 pixels (approximate, varies by font)
        # 1 point height ≈ 1.33 pixels
        
        # Desired Cell Size in Pixels
        CELL_WIDTH_PX = 230
        CELL_HEIGHT_PX = 110

        # Excel Dimensions (Calculated)
        EXCEL_COL_WIDTH = 33      # ~230 pixels wide
        EXCEL_ROW_HEIGHT = 82.5   # ~110 pixels high

        # Calculate Offsets for Centering (Padding)
        # (Container Size - Image Size) / 2
        offset_x_px = (CELL_WIDTH_PX - IMG_WIDTH_PX) / 2
        offset_y_px = (CELL_HEIGHT_PX - IMG_HEIGHT_PX) / 2

        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    match = msid_pattern.search(cell.value)
                    if match:
                        msid_number = match.group(1)
                        print(f"Found MSID: {msid_number} at {cell.coordinate}")

                        # 1. Generate Image
                        img_stream = generate_barcode_image(msid_number)
                        img = ExcelImage(img_stream)
                        
                        # 2. Set Image Size
                        img.width = IMG_WIDTH_PX
                        img.height = IMG_HEIGHT_PX

                        # 3. Determine Target Cell
                        target_row = cell.row + 1
                        target_col = cell.column
                        target_cell = ws.cell(row=target_row, column=target_col)

                        # 4. Resize Row & Column
                        # Set Row Height
                        current_row_h = ws.row_dimensions[target_row].height
                        if current_row_h is None or current_row_h < EXCEL_ROW_HEIGHT:
                             ws.row_dimensions[target_row].height = EXCEL_ROW_HEIGHT

                        # Set Col Width
                        col_letter = get_column_letter(target_col)
                        ws.column_dimensions[col_letter].width = EXCEL_COL_WIDTH

                        # 5. Apply Center Alignment (Text)
                        target_cell.alignment = Alignment(horizontal='center', vertical='center')

                        # 6. Apply Center Anchor (Image)
                        # We use OneCellAnchor with offsets (EMU units) to push the image to the center
                        col_idx = target_col - 1 # 0-indexed
                        row_idx = target_row - 1 # 0-indexed

                        # Define the marker (Top-Left corner + Offset)
                        marker = AnchorMarker(
                            col=col_idx,
                            colOff=pixels_to_EMU(offset_x_px), # Shift Right
                            row=row_idx,
                            rowOff=pixels_to_EMU(offset_y_px)  # Shift Down
                        )

                        # Define the size of the image
                        size = XDRPositiveSize2D(
                            pixels_to_EMU(IMG_WIDTH_PX),
                            pixels_to_EMU(IMG_HEIGHT_PX)
                        )

                        # Create the custom anchor and assign
                        img.anchor = OneCellAnchor(_from=marker, ext=size)
                        ws.add_image(img)

        # Save
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        output_filename = f"{name}_CenteredBarcodes{ext}"
        output_path = os.path.join(get_downloads_folder(), output_filename)

        wb.save(output_path)
        print(f"\nSuccess! File saved to: {output_path}")
        
        if os.name == 'nt':
            os.startfile(get_downloads_folder())
        else:
            os.system(f'open "{get_downloads_folder()}"')

    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    process_msid_sheet()