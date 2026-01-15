import streamlit as st
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
import barcode
from barcode.writer import ImageWriter
import re
import io

# --- PAGE CONFIG ---
st.set_page_config(page_title="MSID Barcode Gen", page_icon="🏷️")
st.title("🏷️ MSID Excel Barcode Generator")
st.markdown("""
**Instructions:**
1. Upload your Excel file containing cells with text like `MSID: 12345`.
2. The app will generate a barcode and place it in the cell **below** the MSID.
""")

# --- HELPER FUNCTIONS ---

def generate_barcode_image(msid_value):
    """Generates a Code128 barcode image in memory."""
    Code128 = barcode.get_barcode_class('code128')
    writer = ImageWriter()
    rv = io.BytesIO()

    # Configure for a barcode that handles scaling well
    code_instance = Code128(msid_value, writer=writer)
    code_instance.write(rv, options={
        'module_width': 0.3,
        'module_height': 8.0,
        'font_size': 6,
        'text_distance': 3.0,
        'quiet_zone': 1.0,
        'write_text': True
    })
    rv.seek(0)
    return rv

def process_excel(file_upload):
    """Reads the uploaded file, adds barcodes, and returns the binary content."""
    # Load workbook from the uploaded file object
    wb = openpyxl.load_workbook(file_upload)
    ws = wb.active

    msid_pattern = re.compile(r'MSID:?\s*(\d+)', re.IGNORECASE)
    
    # Counter for feedback
    count = 0

    # --- SIZE CONFIGURATION (Kept exact from your script) ---
    IMG_WIDTH_PX = 160   
    IMG_HEIGHT_PX = 85   
    CELL_WIDTH_PX = 230
    CELL_HEIGHT_PX = 110
    EXCEL_COL_WIDTH = 33      
    EXCEL_ROW_HEIGHT = 82.5   
    offset_x_px = (CELL_WIDTH_PX - IMG_WIDTH_PX) / 2
    offset_y_px = (CELL_HEIGHT_PX - IMG_HEIGHT_PX) / 2

    # Iterate through rows
    # We convert to list to avoid issues if we modify dimensions while iterating
    for row in list(ws.iter_rows()):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                match = msid_pattern.search(cell.value)
                if match:
                    count += 1
                    msid_number = match.group(1)
                    
                    # 1. Generate Image
                    img_stream = generate_barcode_image(msid_number)
                    img = ExcelImage(img_stream)

                    # 2. Set Image Size
                    img.width = IMG_WIDTH_PX
                    img.height = IMG_HEIGHT_PX

                    # 3. Determine Target Cell
                    target_row = cell.row + 1
                    target_col = cell.column
                    target_cell = ws.cell(row=target_row, column=target_col)

                    # 4. Resize Row & Column
                    current_row_h = ws.row_dimensions[target_row].height
                    if current_row_h is None or current_row_h < EXCEL_ROW_HEIGHT:
                         ws.row_dimensions[target_row].height = EXCEL_ROW_HEIGHT

                    col_letter = get_column_letter(target_col)
                    ws.column_dimensions[col_letter].width = EXCEL_COL_WIDTH

                    # 5. Apply Center Alignment
                    target_cell.alignment = Alignment(horizontal='center', vertical='center')

                    # 6. Apply Center Anchor (Image)
                    col_idx = target_col - 1 
                    row_idx = target_row - 1 

                    marker = AnchorMarker(
                        col=col_idx,
                        colOff=pixels_to_EMU(offset_x_px),
                        row=row_idx,
                        rowOff=pixels_to_EMU(offset_y_px)
                    )

                    size = XDRPositiveSize2D(
                        pixels_to_EMU(IMG_WIDTH_PX),
                        pixels_to_EMU(IMG_HEIGHT_PX)
                    )

                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                    ws.add_image(img)

    # Save to Memory Buffer
    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0) # Rewind buffer to start
    
    return output_buffer, count

# --- MAIN APP LOGIC ---

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    # Button to trigger processing
    if st.button("Generate Barcodes"):
        with st.spinner("Processing... this may take a moment."):
            try:
                # Run the processing function
                processed_data, barcode_count = process_excel(uploaded_file)
                
                if barcode_count > 0:
                    st.success(f"Success! Generated {barcode_count} barcodes.")
                    
                    # Create the download button
                    st.download_button(
                        label="📥 Download Processed Excel",
                        data=processed_data,
                        file_name="MSID_Barcodes.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    st.warning("No 'MSID' patterns found in the uploaded file.")
                    
            except Exception as e:
                st.error(f"An error occurred: {e}")